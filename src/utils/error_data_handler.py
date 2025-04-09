import os
import asyncio
import aiofiles
import openpyxl

from src.logger import AsyncLogger

file_lock = asyncio.Lock()
logger = AsyncLogger()

async def save_bad_discord_token(discord_token: str) -> None:
    config_dir = os.path.join("config", "data", "client")
    bad_token_path = os.path.join(config_dir, "bad_discord_token.txt")
    accounts_path = os.path.join(config_dir, "accounts.xlsx")
    os.makedirs(config_dir, exist_ok=True)
    
    async with file_lock:
        try:
            async with aiofiles.open(bad_token_path, 'a') as file:
                await file.write(f"{discord_token}\n")
            
            if os.path.exists(accounts_path):
                try:
                    wb = openpyxl.load_workbook(accounts_path)
                    ws = wb.active
                    token_col_idx = None
                    
                    for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=False)), 1):
                        if cell.value and cell.value.strip() == 'Discord Token':
                            token_col_idx = idx
                            break
                    
                    if token_col_idx:
                        rows_modified = 0
                        
                        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                            cell = row[token_col_idx-1]
                            cell_value = cell.value
                            if cell_value and cell_value.strip() == discord_token:
                                cell.value = ""
                                rows_modified += 1
                        
                        if rows_modified:
                            wb.save(accounts_path)
                            await logger.logger_msg(
                                f"Cleared bad Discord token from accounts.xlsx",
                                type_msg="info"
                            )
                except Exception as e:
                    await logger.logger_msg(
                        f"Error updating accounts.xlsx: {str(e)}",
                        type_msg="error", method_name="save_bad_discord_token"
                    )
                    
        except Exception as e:
            await logger.logger_msg(
                f"Error processing Discord token: {str(e)}",
                type_msg="error", method_name="save_bad_discord_token"
            )
            
async def save_bad_private_key(private_key: str, wallet_address: str) -> None:
    config_dir = os.path.join("config", "data", "client")
    bad_key_path = os.path.join(config_dir, "bad_private_key.txt")
    accounts_path = os.path.join(config_dir, "accounts.xlsx")
    os.makedirs(config_dir, exist_ok=True)
    
    async with file_lock:
        try:
            async with aiofiles.open(bad_key_path, 'a') as file:
                await file.write(f"{private_key}\n")
            
            if os.path.exists(accounts_path):
                try:
                    wb = openpyxl.load_workbook(accounts_path)
                    ws = wb.active
                    pk_col_idx = None
                    
                    for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=False)), 1):
                        if cell.value and cell.value.strip() == 'Private Key':
                            pk_col_idx = idx
                            break
                    
                    if pk_col_idx:
                        rows_modified = 0
                        
                        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                            cell = row[pk_col_idx-1]
                            cell_value = cell.value
                            if cell_value and cell_value.strip() == private_key:
                                cell.value = ""
                                rows_modified += 1
                        
                        if rows_modified:
                            wb.save(accounts_path)
                            await logger.logger_msg(
                                f"Cleared bad private key from accounts.xlsx",
                                type_msg="info", address=wallet_address
                            )
                except Exception as e:
                    await logger.logger_msg(
                        f"Error updating accounts.xlsx: {str(e)}", type_msg="error", 
                        address=wallet_address, method_name="save_bad_private_key"
                    )
                    
        except Exception as e:
            await logger.logger_msg(
                f"Error processing private key: {str(e)}", type_msg="error", 
                address=wallet_address, method_name="save_bad_private_key"
            )

async def save_bad_twitter_token(twitter_token: str, wallet_address: str = None) -> None:
    config_dir = os.path.join("config", "data", "client")
    bad_token_path = os.path.join(config_dir, "bad_twitter_token.txt")
    accounts_path = os.path.join(config_dir, "accounts.xlsx")
    os.makedirs(config_dir, exist_ok=True)
    
    async with file_lock:
        try:
            async with aiofiles.open(bad_token_path, 'a') as file:
                await file.write(f"{twitter_token}\n")
            
            if os.path.exists(accounts_path):
                try:
                    wb = openpyxl.load_workbook(accounts_path)
                    ws = wb.active
                    token_col_idx = None
                    
                    for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=False)), 1):
                        if cell.value and cell.value.strip() == 'Twitter Token':
                            token_col_idx = idx
                            break
                    
                    if token_col_idx:
                        rows_modified = 0
                        
                        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                            cell = row[token_col_idx-1]
                            cell_value = cell.value
                            if cell_value and cell_value.strip() == twitter_token:
                                cell.value = ""
                                rows_modified += 1
                        
                        if rows_modified:
                            wb.save(accounts_path)
                            await logger.logger_msg(
                                f"Cleared bad Twitter token from accounts.xlsx", 
                                type_msg="info", address=wallet_address
                            )
                except Exception as e:
                    await logger.logger_msg(
                        f"Error updating accounts.xlsx: {str(e)}", type_msg="error", 
                        address=wallet_address, method_name="save_bad_twitter_token"
                    )
                    
        except Exception as e:
            await logger.logger_msg(
                f"Error processing Twitter token: {str(e)}", type_msg="error", 
                address=wallet_address, method_name="save_bad_twitter_token"
            )
            
async def check_twitter_error_for_invalid_token(error_message: str, twitter_token: str, wallet_address: str) -> bool:
    error_str = str(error_message).lower()
    auth_error_keywords = [
        "auth", "token", "unauthorized", "invalid token", "expired", 
        "authentication", "login", "credentials", "401", "403"
    ]
    
    if any(keyword in error_str for keyword in auth_error_keywords):
        await logger.logger_msg(
            f"Detected invalid Twitter token", type_msg="warning", 
            address=wallet_address, method_name="check_twitter_error_for_invalid_token"
        )
        await save_bad_twitter_token(twitter_token, wallet_address)
        return True
    
    return False