import os
import asyncio
import aiofiles
import openpyxl
from typing import Union

from src.logger import AsyncLogger


# Global constants
CONFIG_DIR = os.path.join("config", "data", "client")
ACCOUNTS_PATH = os.path.join(CONFIG_DIR, "accounts.xlsx")
BAD_DISCORD_TOKEN_PATH = os.path.join(CONFIG_DIR, "bad_discord_token.txt")
BAD_PRIVATE_KEY_PATH = os.path.join(CONFIG_DIR, "bad_private_key.txt")
BAD_TWITTER_TOKEN_PATH = os.path.join(CONFIG_DIR, "bad_twitter_token.txt")

# Constants for column names
COL_DISCORD_TOKEN = 'Discord Token'
COL_TWITTER_TOKEN = 'Twitter Token'
COL_PRIVATE_KEY = 'Private Key'
COL_RECONNECT_DISCORD = 'Reconnect Discord'
COL_RECONNECT_TWITTER = 'Reconnect Twitter'
COL_NATIVE_BALANCE = 'Native Balance $STT'
COL_ADDRESS = 'Address'

# File lock for safe file operations
file_lock = asyncio.Lock()
logger = AsyncLogger()


async def is_string_in_file(file_path: str, search_string: str) -> bool:
    """
    Checks if the string is in the file
    """
    if not os.path.exists(file_path):
        return False
        
    async with aiofiles.open(file_path, 'r') as file:
        content = await file.read()
        return search_string in content.splitlines()


def get_excel_column_mapping(worksheet) -> dict[str, int]:
    """
    Creates a mapping of column headers to their indices
    """
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=False))
    col_map = {}
    
    for idx, cell in enumerate(header_row):
        if cell.value:
            col_map[cell.value.strip()] = idx
    
    return col_map


async def save_to_bad_token_file(file_path: str, token: str) -> None:
    """
    Saves the token to the file of invalid tokens if it is not already there
    """
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    if not await is_string_in_file(file_path, token):
        async with aiofiles.open(file_path, 'a') as file:
            await file.write(f"{token}\n")


async def update_excel_for_bad_token(
    token: str, 
    token_column_name: str, 
    reconnect_column_name: str,
    wallet_address: str | None = None
) -> None:
    """
    Updates the Excel file when an invalid token is detected
    """
    if not os.path.exists(ACCOUNTS_PATH):
        return
        
    try:
        wb = openpyxl.load_workbook(ACCOUNTS_PATH)
        ws = wb.active
        
        col_map = get_excel_column_mapping(ws)
        
        token_col_idx = col_map.get(token_column_name)
        if token_col_idx is None:
            return
        
        reconnect_col_idx = col_map.get(reconnect_column_name)
        if reconnect_col_idx is None:
            return
        
        rows_modified = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
            cell = row[token_col_idx]
            cell_value = cell.value
            
            if cell_value and str(cell_value).strip() == token:
                cell.value = ""
                ws.cell(row=row_idx, column=reconnect_col_idx + 1).value = True
                rows_modified += 1
        
        if rows_modified > 0:
            wb.save(ACCOUNTS_PATH)
            
            log_msg = f"Cleared bad {token_column_name} and set {reconnect_column_name} flag to True"
            log_kwargs = {"type_msg": "info"}
            
            if wallet_address:
                log_kwargs["address"] = wallet_address
                
            await logger.logger_msg(log_msg, **log_kwargs)
            
    except Exception as e:
        method_name = f"update_excel_for_bad_token({token_column_name})"
        log_kwargs = {"type_msg": "error", "method_name": method_name}
        
        if wallet_address:
            log_kwargs["address"] = wallet_address
            
        await logger.logger_msg(f"Error updating accounts.xlsx: {str(e)}", **log_kwargs)


async def save_bad_discord_token(discord_token: str, wallet_address: str) -> None:
    """
    Handles an invalid Discord token
    """
    async with file_lock:
        try:
            await save_to_bad_token_file(BAD_DISCORD_TOKEN_PATH, discord_token)
            
            await update_excel_for_bad_token(
                token=discord_token,
                token_column_name=COL_DISCORD_TOKEN,
                reconnect_column_name=COL_RECONNECT_DISCORD,
                wallet_address=wallet_address
            )
                
        except Exception as e:
            await logger.logger_msg(
                f"Error processing Discord token: {str(e)}",
                type_msg="error", 
                address=wallet_address,
                method_name="save_bad_discord_token"
            )
            

async def save_bad_private_key(private_key: str, wallet_address: str) -> None:
    """
    Handles an invalid private key
    """
    async with file_lock:
        try:
            await save_to_bad_token_file(BAD_PRIVATE_KEY_PATH, private_key)
            
            if os.path.exists(ACCOUNTS_PATH):
                try:
                    wb = openpyxl.load_workbook(ACCOUNTS_PATH)
                    ws = wb.active
                    
                    col_map = get_excel_column_mapping(ws)
                    pk_col_idx = col_map.get(COL_PRIVATE_KEY)
                    
                    if pk_col_idx is not None:
                        rows_modified = 0
                        
                        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                            cell = row[pk_col_idx]
                            cell_value = cell.value
                            
                            if cell_value and str(cell_value).strip() == private_key:
                                cell.value = ""
                                rows_modified += 1
                        
                        if rows_modified:
                            wb.save(ACCOUNTS_PATH)
                            await logger.logger_msg(
                                f"Cleared bad private key from accounts.xlsx",
                                type_msg="info", 
                                address=wallet_address
                            )
                except Exception as e:
                    await logger.logger_msg(
                        f"Error updating accounts.xlsx: {str(e)}", 
                        type_msg="error", 
                        address=wallet_address, 
                        method_name="save_bad_private_key"
                    )
                    
        except Exception as e:
            await logger.logger_msg(
                f"Error processing private key: {str(e)}", 
                type_msg="error", 
                address=wallet_address, 
                method_name="save_bad_private_key"
            )


async def save_bad_twitter_token(twitter_token: str, wallet_address: str = None) -> None:
    """
    Handles an invalid Twitter token
    """
    async with file_lock:
        try:
            await save_to_bad_token_file(BAD_TWITTER_TOKEN_PATH, twitter_token)
            
            await update_excel_for_bad_token(
                token=twitter_token,
                token_column_name=COL_TWITTER_TOKEN,
                reconnect_column_name=COL_RECONNECT_TWITTER,
                wallet_address=wallet_address
            )
                
        except Exception as e:
            await logger.logger_msg(
                f"Error processing Twitter token: {str(e)}", 
                type_msg="error", 
                address=wallet_address, 
                method_name="save_bad_twitter_token"
            )


async def check_twitter_error_for_invalid_token(
    error_message: Union[str, dict], 
    twitter_token: str, 
    wallet_address: str
) -> bool:
    """
    Checks if the error indicates an invalid Twitter token
    """
    if isinstance(error_message, dict):
        error_code = error_message.get('error_code')
        if error_code in (32, 89, 135):
            await logger.logger_msg(
                f"Detected invalid Twitter token (error code {error_code})", 
                type_msg="warning", 
                address=wallet_address, 
                method_name="check_twitter_error_for_invalid_token"
            )
            await save_bad_twitter_token(twitter_token, wallet_address)
            return True

    error_str = str(error_message).lower()
    
    if any(code in error_str for code in ["401", "403", "status: 401", "status: 403"]):
        if any(auth_term in error_str for auth_term in ["auth", "token", "login", "credentials"]):
            await logger.logger_msg(
                f"Detected authentication error in Twitter API", 
                type_msg="warning", 
                address=wallet_address, 
                method_name="check_twitter_error_for_invalid_token"
            )
            await save_bad_twitter_token(twitter_token, wallet_address)
            return True
    
    return False


async def clear_token_after_successful_connection(
    token: str, 
    token_column_name: str, 
    wallet_address: str | None = None
) -> None:
    """
    Clears the token from the Excel file after successful connection to a social network
    """
    if not os.path.exists(ACCOUNTS_PATH):
        return
        
    try:
        wb = openpyxl.load_workbook(ACCOUNTS_PATH)
        ws = wb.active
        
        col_map = get_excel_column_mapping(ws)
        
        token_col_idx = col_map.get(token_column_name)
        if token_col_idx is None:
            return
        
        rows_modified = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
            cell = row[token_col_idx]
            cell_value = cell.value
            
            if cell_value and str(cell_value).strip() == token:
                cell.value = ""
                rows_modified += 1
        
        if rows_modified > 0:
            wb.save(ACCOUNTS_PATH)
            
            log_msg = f"Cleared {token_column_name} after successful connection"
            log_kwargs = {"type_msg": "info"}
            
            if wallet_address:
                log_kwargs["address"] = wallet_address
                
            await logger.logger_msg(log_msg, **log_kwargs)
            
    except Exception as e:
        method_name = f"clear_token_after_successful_connection({token_column_name})"
        log_kwargs = {"type_msg": "error", "method_name": method_name}
        
        if wallet_address:
            log_kwargs["address"] = wallet_address
            
        await logger.logger_msg(f"Error updating accounts.xlsx: {str(e)}", **log_kwargs)


async def update_native_balance_in_excel(wallet_address: str, balance: str, private_key: str = None) -> None:
    """
    Updates the wallet balance value in an Excel file
    """
    if not os.path.exists(ACCOUNTS_PATH):
        return
        
    try:
        async with file_lock:
            wb = openpyxl.load_workbook(ACCOUNTS_PATH)
            ws = wb.active
            
            col_map = get_excel_column_mapping(ws)
            
            balance_col_idx = col_map.get(COL_NATIVE_BALANCE)
            if balance_col_idx is None:
                return
                
            address_col_idx = col_map.get(COL_ADDRESS)
            private_key_col_idx = col_map.get(COL_PRIVATE_KEY)
            
            wallet_address_found = False
            
            if address_col_idx is not None:
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                    cell = row[address_col_idx]
                    if cell.value and str(cell.value).lower() == wallet_address.lower():
                        ws.cell(row=row_idx, column=balance_col_idx + 1).value = balance
                        wallet_address_found = True
                        break
            
            if not wallet_address_found and private_key and private_key_col_idx is not None:
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
                    cell = row[private_key_col_idx]
                    if cell.value and str(cell.value).strip() == private_key:
                        ws.cell(row=row_idx, column=balance_col_idx + 1).value = balance
                        
                        if address_col_idx is not None:
                            ws.cell(row=row_idx, column=address_col_idx + 1).value = wallet_address
                            
                        wallet_address_found = True
                        break
                        
            if wallet_address_found:
                wb.save(ACCOUNTS_PATH)
                await logger.logger_msg(
                    f"Updated native balance for {wallet_address} to {balance}",
                    type_msg="info", 
                    address=wallet_address
                )
            else:
                await logger.logger_msg(
                    f"Wallet address or private key not found in Excel",
                    type_msg="warning", 
                    address=wallet_address, 
                    method_name="update_native_balance_in_excel"
                )
                
    except Exception as e:
        await logger.logger_msg(
            f"Error updating native balance in Excel: {str(e)}", 
            type_msg="error", 
            address=wallet_address, 
            method_name="update_native_balance_in_excel"
        )